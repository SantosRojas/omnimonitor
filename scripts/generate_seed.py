#!/usr/bin/env python3
"""
Generate seed_data.rs from data.xlsx

Usage:
    python scripts/generate_seed.py

Reads data.xlsx from the project root, translates English signal descriptions
to short Spanish display names, and generates server/src/infrastructure/seed_data.rs.
"""

import re
import openpyxl
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "data.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "server" / "src" / "infrastructure" / "seed_data.rs"

# ═══════════════════════════════════════════════════
#  Spanish display name translations
#  If a description isn't in this map, the script
#  falls back to the English description as-is.
# ═══════════════════════════════════════════════════

TRANSLATIONS = {
    # ── Pressure / Flow ──
    "measured arterial pressure (filtered)": "Presión arterial (filtrada)",
    "measured arterial pressure": "Presión arterial",
    "measured venous pressure (filtered)": "Presión venosa (filtrada)",
    "measured venous pressure": "Presión venosa",
    "measured filter pressure (filtered)": "Presión de filtro (filtrada)",
    "measured filter pressure": "Presión de filtro",
    "measured effluent pressure (filtered)": "Presión de efluente (filtrada)",
    "measured effluent pressure": "Presión de efluente",
    "measured solution pressure (filtered)": "Presión de solución (filtrada)",
    "measured solution pressure": "Presión de solución",
    "measured arterial pressure (filtered)": "Presión arterial (filtrada)",
    "arterial pressure window maximum (actual)": "Presión arterial máx.",
    "arterial pressure window minimum (actual)": "Presión arterial mín.",
    "upper limit of venous pressure": "Límite superior presión venosa",
    "lower limit of venous pressure": "Límite inferior presión venosa",
    "upper limit of filter pressure": "Límite superior presión filtro",
    "lower limit of filter pressure": "Límite inferior presión filtro",
    "actual filter pressure drop": "Caída de presión en filtro",
    "actual transmembrane pressure": "Presión transmembrana",

    # ── Weight / Bags ──
    "measured bag weight (filtered)": "Peso de bolsa (filtrado)",
    "measured bag weight": "Peso de bolsa",
    "bag remaining time until empty": "Tiempo restante de bolsa",
    "middle bag numbers set by user": "N° bolsas media (usuario)",
    "middle bag numbers estimated by machine": "N° bolsas media (estimado)",

    # ── Flow rates / Pumps ──
    "blood side blood pump actual flow": "Flujo bomba de sangre",
    "blood side blood pump actual rotation": "Rotación bomba de sangre",
    "blood side top pump actual flow": "Flujo bomba superior (sangre)",
    "blood side top pump actual rotation": "Rotación bomba superior (sangre)",
    "fluid side middle pump actual flow": "Flujo bomba media (fluido)",
    "fluid side middle pump actual rotation": "Rotación bomba media (fluido)",
    "fluid side bottom pump actual flow": "Flujo bomba inferior (fluido)",
    "fluid side bottom pump actual rotation": "Rotación bomba inferior (fluido)",
    "fluid side top pump actual flow": "Flujo bomba superior (fluido)",
    "fluid side top pump actual rotation": "Rotación bomba superior (fluido)",
    "syringe pump actual flow": "Flujo bomba de jeringa",
    "syringe pump actual rotation": "Rotación bomba de jeringa",
    "blood side blood pump tube constant": "Constante tubular bomba sangre",
    "filter (dialyser) factor measured by machine": "Factor de filtro (dializador)",

    # ── Volume / Removal ──
    "actual net fluid removal volume": "Volumen neto de remoción",
    "blood side accumulated blood volume": "Volumen acumulado de sangre",
    "fluid side total infused volume": "Volumen total infundido",
    "fluid side total removed volume": "Volumen total removido",

    # ── Temperature ──
    "measured output fluid temperature (filtered)": "Temperatura de salida (filtrada)",
    "measured output fluid temperature": "Temperatura de salida",
    "measured warmer plate temperature (filtered)": "Temperatura de placa (filtrada)",
    "measured warmer plate temperature": "Temperatura de placa",

    # ── Air detectors ──
    "accumulated air bubble volume by citrate safety air detector (micro litre)": "Burbujas acumuladas (detector citrato)",
    "continuous air volume by citrate safety air detector (micro litre)": "Aire continuo (detector citrato)",
    "accumulated air bubble volume by venous safety air detector (micro litre)": "Burbujas acumuladas (detector venoso)",
    "continuous air volume by venous safety air detector (micro litre)": "Aire continuo (detector venoso)",

    # ── Time ──
    "remaining time until the periodic test (24h)": "Tiempo restante prueba periódica",
    "kit remaining time until kit change": "Tiempo restante cambio de kit",
    "syringe remaining time until empty": "Tiempo restante de jeringa",

    # ── Battery ──
    "battery voltage level": "Nivel de voltaje batería",

    # ── State machines ──
    "control sw main state": "Estado ppal. SW control",
    "control sw sub state": "Subestado SW control",
    "protective sw main state": "Estado ppal. SW protección",
    "protective sw sub state": "Subestado SW protección",
    "graphical user interface sw main state": "Estado ppal. interfaz gráfica",
    "graphical user interface sw sub state": "Subestado interfaz gráfica",
    "id of the selected anticoagulation mode": "Modo de anticoagulación",
    "id of the selected substitution mode": "Modo de sustitución",
    "id of the selected therapy type": "Tipo de terapia",

    # ── Configuration / Constants ──
    "set citrate concentration": "Concentración de citrato",
    "set calcium concentration": "Concentración de calcio",
    "set blood flow": "Flujo de sangre configurado",
    "set fluid flow": "Flujo de fluido configurado",
    "set patient temperature": "Temperatura del paciente configurada",
    "set fluid temperature": "Temperatura del fluido configurada",
    "patient target weight": "Peso objetivo del paciente",
    "set ultrafiltration rate": "Tasa de ultrafiltración",
    "set heparin rate": "Tasa de heparina",
    "set dialysis flow rate": "Flujo de diálisis",
    "set substitution flow rate": "Flujo de sustitución",
    "patient fluid removal goal": "Meta de remoción de fluido",
    "patient fluid removal rate": "Tasa de remoción de fluido",
    "anticoagulation syringe volume set": "Volumen jeringa anticoagulación",
    "treatment time prescribed": "Tiempo de tratamiento prescrito",
    "treatment time elapsed": "Tiempo de tratamiento transcurrido",
    "treatment time remaining": "Tiempo de tratamiento restante",
    "total heparin dose": "Dosis total de heparina",
    "heparin infusion rate": "Tasa de infusión heparina",
    "citrate infusion rate": "Tasa de infusión citrato",
    "calcium infusion rate": "Tasa de infusión calcio",

    # ── Machine info ──
    "Serial number": "N° de serie",
    "kit type": "Tipo de kit",
    "patient name": "Nombre del paciente",
    "syringe type": "Tipo de jeringa",
    "software version": "Versión de software",
    "hardware version": "Versión de hardware",

    # ── Alarm / Event ──
    "alarm active": "Alarma activa",
    "alarm type": "Tipo de alarma",
    "alarm code": "Código de alarma",
    "event code": "Código de evento",
    "event text": "Texto de evento",

    # ── Fluid composition ──
    "fluid sodium concentration": "Concentración de sodio",
    "fluid potassium concentration": "Concentración de potasio",
    "fluid calcium concentration": "Concentración de calcio",
    "fluid magnesium concentration": "Concentración de magnesio",
    "fluid chloride concentration": "Concentración de cloruro",
    "fluid bicarbonate concentration": "Concentración de bicarbonato",
    "fluid glucose concentration": "Concentración de glucosa",
    "fluid lactate concentration": "Concentración de lactato",
    "fluid ph value": "pH del fluido",

    # ── Pressure limits ──
    "upper limit of arterial pressure": "Límite superior presión arterial",
    "lower limit of arterial pressure": "Límite inferior presión arterial",
    "upper limit of effluent pressure": "Límite superior presión efluente",
    "lower limit of effluent pressure": "Límite inferior presión efluente",

    # ── Access ──
    "access pressure": "Presión de acceso",
    "access flow": "Flujo de acceso",
    "return pressure": "Presión de retorno",

    # ── Accumulated volumes ──
    "accumulated (treated) blood volume": "Vol. sangre tratado acum.",
    "accumulated calcium volume (during treatment)": "Vol. calcio acumulado",
    "accumulated citrate volume": "Vol. citrato acumulado",
    "accumulated dialysate volume": "Vol. dializado acumulado",
    "accumulated effluent volume": "Vol. efluente acumulado",
    "accumulated heparin volume": "Vol. heparina acumulado",
    "accumulated net fluid removal volume": "Remoción neta acumulada",
    "accumulated plasma substitution volume": "Vol. sustitución plasma acum.",
    "accumulated plasma volume": "Vol. plasma acumulado",
    "accumulated post substitution volume": "Vol. sust. post acumulado",
    "accumulated pre substitution volume": "Vol. sust. pre acumulado",

    # ── Actual flow / delivery ──
    "actual blood flow": "Flujo de sangre actual",
    "actual calcium flow added by substitution": "Flujo de calcio (sustitución)",
    "actual calcium ratio value": "Relación de calcio actual",
    "actual citrate ratio value": "Relación de citrato actual",
    "actual delivered blood volume": "Vol. sangre administrado",
    "actual delivered heparin bolus volume": "Bolo heparina administrado",
    "actual net fluid removal flow": "Flujo remoción neta actual",

    # ── Set values (user configuration) ──
    "blood flow set by the user": "Flujo de sangre configurado",
    "dialysate flow set by the user": "Flujo de diálisis configurado",
    "fluid temperature set by the user": "Temp. fluido configurada",
    "net fluid removal flow set by the user": "Flujo remoción neta config.",
    "net fluid removal set by the user": "Remoción neta configurada",
    "calcium ratio set by the user": "Relación de calcio config.",
    "citrate ratio set by the user": "Relación de citrato config.",
    "heparin bolus volume set by the user": "Bolo heparina configurado",
    "heparin flow set by the user": "Flujo heparina configurado",
    "therapy time set by the user": "Tiempo de terapia config.",
    "patient weight set by the user": "Peso paciente configurado",
    "patient hematocrit value set by the user": "Hematocrito configurado",
    "plasma filtration ratio set by the user": "Relación filtración plasma config.",
    "plasma sbstitution volume set by the user": "Vol. sustitución plasma config.",
    "post-substitution flow set by the user": "Flujo sust. post configurado",
    "pre-substitution flow set by the user": "Flujo sust. pre configurado",
    "syringe filled volume set by the user": "Vol. jeringa configurado",

    # ── Pressure limits (user-set) ──
    "max arterial pressure window limit set by the user": "Ventana presión arterial máx.",
    "min arterial pressure window limit set by the user": "Ventana presión arterial mín.",
    "max venous pressure window limit set by the user": "Ventana presión venosa máx.",
    "min venous pressure window limit set by the user": "Ventana presión venosa mín.",
    "max filter pressure window limit set by the user": "Ventana presión filtro máx.",
    "min filter pressure window limit set by the user": "Ventana presión filtro mín.",
    "max filter pressure limit set by the user": "Presión filtro máx. config.",
    "max filter pressure drop limit set by the user": "Caída presión filtro máx.",
    "max transmembrane pressure limit set by the user": "Presión transmembrana máx.",
    "maximum arterial pressure limit set by the user": "Presión arterial máx. config.",
    "minimum arterial pressure limit set by the user": "Presión arterial mín. config.",
    "minimum effluent pressure limit set by the user": "Presión efluente mín. config.",
    "maximum venous pressure limit set by the user": "Presión venosa máx. config.",

    # ── Battery ──
    "measured battery current": "Corriente de batería",
    "measured battery voltage": "Voltaje de batería",

    # ── Derived / Calculated ──
    "average target renal dose based on set values": "Dosis renal objetivo promedio",
    "current target renal dose based on set values": "Dosis renal objetivo actual",
    "renal dose measured by machine": "Dosis renal medida",
    "fluid/blood filtration ratio calculated by the machine": "Relación filtración fluido/sangre",
    "percentage of calcium flow added by substitution": "% flujo calcio en sustitución",
    "required total calcium flow": "Flujo total de calcio requerido",
    "summed substitution flow in post-post dilution mode": "Flujo sust. total (post-post)",
    "estimated net fluid removal volume": "Remoción neta estimada",
    "elapsed therapy time": "Tiempo de terapia transcurrido",

    # ── Syringe ──
    "syringe pump calcium stop time": "Tiempo paro calcio jeringa",
    "syringe remaining volume": "Vol. restante jeringa",

}

DEFAULT_TRANSLATION = "—"  # fallback for untranslated descriptions

# ═══════════════════════════════════════════════════
#  Spanish translations for equivalence display names
#  (state machine values from the equivalences sheet)
# ═══════════════════════════════════════════════════

EQUIV_TRANSLATIONS = {
    # ── Top-level states ──
    "Preparation": "Preparación",
    "Connect patient": "Conectar paciente",
    "Therapy": "Terapia",
    "End of therapy": "Fin de terapia",
    "Temporarily disconnect": "Desconexión temporal",
    "Change kit": "Cambiar kit",
    "Change therapy": "Cambiar terapia",
    "Change dilution": "Cambiar dilución",
    "Change anticoag": "Cambiar anticoagulación",

    # ── Preparation sub-states ──
    "Preparation start screen": "Preparación - inicio",
    "Preparation history view": "Preparación - historial",
    "Preparation previous kit removal": "Preparación - retirar kit anterior",
    "Preparation previous kit removal stop": "Preparación - retirar kit (paro)",
    "Preparation previous kit removal unloading": "Preparación - retirar kit (descarga)",
    "Preparation scan kit": "Preparación - escanear kit",
    "Preparation select therapy": "Preparación - seleccionar terapia",
    "Preparation device test": "Preparación - prueba del equipo",
    "Preparation device test finished": "Preparación - prueba finalizada",
    "Preparation install kit": "Preparación - instalar kit",
    "Preparation install kit stop": "Preparación - instalar kit (paro)",
    "Preparation install kit loading": "Preparación - instalar kit (carga)",
    "Preparation install kit unloading": "Preparación - instalar kit (descarga)",
    "Preparation install kit ready": "Preparación - kit instalado",
    "Preparation install bags": "Preparación - instalar bolsas",
    "Preparation install bags ready": "Preparación - bolsas instaladas",
    "Preparation priming": "Preparación - cebado",
    "Preparation priming stop": "Preparación - cebado (paro)",
    "Preparation priming blood side filling start": "Cebado - inicio llenado sangre",
    "Preparation priming blood side art line filling": "Cebado - llenado línea arterial",
    "Preparation priming blood side ven line filling": "Cebado - llenado línea venosa",
    "Preparation priming blood side fpc level adjustment": "Cebado - ajuste FPC",
    "Preparation priming blood side vpc level adjustment": "Cebado - ajuste VPC",
    "Preparation priming blood side filled": "Cebado - lado sangre lleno",
    "Preparation priming fluid side filling start": "Cebado - inicio llenado fluido",
    "Preparation priming fluid side dial line filling": "Cebado - llenado línea dial.",
    "Preparation priming fluid side ef line filling": "Cebado - llenado línea EF",
    "Preparation priming fluid side su g pre line filling": "Cebado - llenado SU G pre",
    "Preparation priming fluid side su g post line filling": "Cebado - llenado SU G post",
    "Preparation priming fluid side su p main line filling": "Cebado - llenado SU P",
    "Preparation priming fluid side plasma line filling": "Cebado - llenado plasma",
    "Preparation priming fluid side flushing line filling": "Cebado - llenado flushing",
    "Preparation priming fluid side spc level adjustment": "Cebado - ajuste SPC",
    "Preparation priming fluid side bypass line filling": "Cebado - llenado bypass",
    "Preparation priming fluid side seq ef line filling": "Cebado - llenado SEQ EF",
    "Preparation priming fluid side epc level adjustment": "Cebado - ajuste EPC",
    "Preparation priming fluid side filled": "Cebado - lado fluido lleno",
    "Preparation priming dialyser degassing": "Cebado - desgasificado",
    "Preparation priming required rinsing": "Cebado - enjuague requerido",
    "Preparation priming final test": "Cebado - prueba final",
    "Preparation priming final rinsing": "Cebado - enjuague final",
    "Preparation priming rinsing finished": "Cebado - enjuague completado",
    "Preparation priming ready": "Cebado - listo",
    "Preparation ready for therapy": "Preparación - lista para terapia",
    "Preparation optional rinsing": "Preparación - enjuague opcional",
    "Preparation optional rinsing ready": "Preparación - enjuague opcional listo",
    "Preparation recirculation": "Preparación - recirculación",
    "Preparation confirm": "Preparación - confirmar",
    "Preparation cancel": "Preparación - cancelar",
    "Preparation cancel stop": "Preparación - cancelar (paro)",
    "Preparation cancel unloading": "Preparación - cancelar (descarga)",
    "Preparation leaving": "Preparación - saliendo",

    # ── Connect patient sub-states ──
    "Connect patient stop": "Conectar paciente (paro)",
    "Connect patient running ac stop": "Conectar paciente - AC paro",
    "Connect patient running ac running": "Conectar paciente - AC activo",
    "Connect patient ready": "Conectar paciente - listo",

    # ── Therapy sub-states ──
    "Therapy main": "Terapia - principal",
    "Therapy blood stop": "Terapia - paro sangre",
    "Therapy patient care blood circ ac stop": "Terapia - cuidado pac. AC paro",
    "Therapy patient care blood circ ac running": "Terapia - cuidado pac. AC activo",
    "Therapy blood circ ac stop": "Terapia - circulación AC paro",
    "Therapy blood circ ac running": "Terapia - circulación AC activo",
    "Therapy running ac stop": "Terapia - AC paro",
    "Therapy running ac running": "Terapia - AC activo",
    "Therapy periodic test": "Terapia - prueba periódica",

    # ── End of therapy sub-states ──
    "End of therapy return blood": "Fin de terapia - retornar sangre",
    "End of therapy return blood stop": "Fin de terapia - retornar sangre (paro)",
    "End of therapy return blood running": "Fin de terapia - retornando sangre",
    "End of therapy return blood ready": "Fin de terapia - retorno listo",
    "End of therapy disconnect all": "Fin de terapia - desconectar todo",
    "End of therapy disconnect all stop": "Fin de terapia - desconectar (paro)",
    "End of therapy disconnect all unloading": "Fin de terapia - desconectar (descarga)",
    "End of therapy leaving": "Fin de terapia - saliendo",

    # ── Temporarily disconnect sub-states ──
    "Temporarily disconnect patient return blood": "Desconexión temp. - retornar sangre",
    "Temporarily disconnect patient return blood stop": "Desconexión temp. - retorno (paro)",
    "Temporarily disconnect patient return blood running": "Desconexión temp. - retornando sangre",
    "Temporarily disconnect patient return blood ready": "Desconexión temp. - retorno listo",
    "Temporarily disconnect patient disconnect pat": "Desconexión temp. - desconectar pac.",
    "Temporarily disconnect patient recirculation": "Desconexión temp. - recirculación",
    "Temporarily disconnect patient recirculation stop": "Desconexión temp. - recirculación (paro)",
    "Temporarily disconnect patient recirculation running": "Desconexión temp. - recirculando",
    "Temporarily disconnect patient confirm": "Desconexión temp. - confirmar",
    "Temporarily disconnect patient connect patient": "Desconexión temp. - conectar pac.",
    "Temporarily disconnect patient connect patient stop": "Desconexión temp. - conectar (paro)",
    "Temporarily disconnect patient connect patient running ac stop": "Desconexión temp. - AC paro",
    "Temporarily disconnect patient connect patient running ac running": "Desconexión temp. - AC activo",
    "Temporarily disconnect patient connect patient ready": "Desconexión temp. - conexión lista",

    # ── Change therapy sub-states ──
    "Change therapy select therapy": "Cambiar terapia - seleccionar",
    "Change therapy connect lines": "Cambiar terapia - conectar líneas",
    "Change therapy confirm": "Cambiar terapia - confirmar",
    "Change therapy blood stop": "Cambiar terapia - paro sangre",
    "Change therapy blood circ ac stop": "Cambiar terapia - circulación AC paro",
    "Change therapy blood circ ac running": "Cambiar terapia - circulación AC activo",

    # ── Change dilution sub-states ──
    "Change dilution select dilution": "Cambiar dilución - seleccionar",
    "Change dilution connect lines": "Cambiar dilución - conectar líneas",
    "Change dilution confirm": "Cambiar dilución - confirmar",
    "Change dilution blood stop": "Cambiar dilución - paro sangre",
    "Change dilution blood circ ac stop": "Cambiar dilución - circulación AC paro",
    "Change dilution blood circ ac running": "Cambiar dilución - circulación AC activo",

    # ── Change kit sub-states ──
    "Change kit return blood": "Cambiar kit - retornar sangre",
    "Change kit return blood stop": "Cambiar kit - retorno (paro)",
    "Change kit return blood running": "Cambiar kit - retornando sangre",
    "Change kit return blood ready": "Cambiar kit - retorno listo",
    "Change kit disconnect kit": "Cambiar kit - desconectar kit",
    "Change kit disconnect kit stop": "Cambiar kit - desconectar (paro)",
    "Change kit disconnect kit unloading": "Cambiar kit - desconectar (descarga)",

    # ── Change anticoag sub-states ──
    "Change anticoag select anticoag": "Cambiar anticoag. - seleccionar",
    "Change anticoag unload syringe": "Cambiar anticoag. - retirar jeringa",
    "Change anticoag change bags": "Cambiar anticoag. - cambiar bolsas",
    "Change anticoag confirm": "Cambiar anticoag. - confirmar",
    "Change anticoag blood stop": "Cambiar anticoag. - paro sangre",
    "Change anticoag blood circ ac stop": "Cambiar anticoag. - circulación AC paro",

    # ── Mode values ──
    "Not selected": "No seleccionado",
    "None": "Ninguno",
    "Citrate calcium": "Citrato calcio",
    "Heparint": "Heparina",
    "Post": "Post",
    "Pre post": "Pre + Post",
    "Pre": "Pre",
    "Post post": "Post + Post",

    # ── Therapy modes ──
    "No valid therapy selected": "Sin terapia seleccionada",
    "Slow Continuous Ultrafiltration": "UF lenta continua (SCUF)",
    "Continuous Veno-Venous Hemofiltration": "Hemofiltración VVC (CVVH)",
    "Continuous Veno-Venous Hemodialysis": "Hemodiálisis VVC (CVVHD)",
    "Continuous Veno-Venous Hemodiafiltration": "Hemodiafiltración VVC (CVVHDF)",
    "Hemoperfusion": "Hemoperfusión (HP)",
    "Therapeutic Plasma Exchange": "Recambio plasmático (TPE)",
    "Cascade Filtration": "Filtración en cascada",
    "Plasma adsorption/perfusion": "Adsorción/perfusión plasmática",
    "Low Blood Volume Slow Continuous Ultrafiltration": "UF lenta continua BVS (SCUF)",
    "Low Blood Volume Continuous Veno-Venous Hemofiltration": "Hemofiltración VVC BVS (CVVH)",
    "Low Blood Volume Continuous Veno-Venous Hemodialysis": "Hemodiálisis VVC BVS (CVVHD)",
    "Low Blood Volume Continuous Veno-Venous Hemodiafiltration": "Hemodiafiltración VVC BVS (CVVHDF)",
}

DEFAULT_TRANSLATION = "—"  # fallback for untranslated descriptions


def translate(description: str | None) -> str:
    """Convert an English description to a short Spanish display name."""
    if not description:
        return DEFAULT_TRANSLATION
    desc = description.strip()
    return TRANSLATIONS.get(desc, desc)  # fallback to English if not translated


def translate_equiv(name: str | None) -> str:
    """Convert an English equivalence display name to Spanish."""
    if not name:
        return ""
    return EQUIV_TRANSLATIONS.get(name.strip(), name.strip())


def to_rust_string(s: str | None) -> str:
    """Escape a string for use as a Rust string literal."""
    if s is None:
        return "None"
    escaped = (
        s.replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\n", "\\n")
        .replace("\t", "\\t")
    )
    return f'Some("{escaped}")'


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # ── Read signals ──────────────────────────────
    ws_signals = wb["signals"]
    # internal_name -> (display_name_es, display_name_en, unit)
    signals: dict[str, tuple[str, str, str | None]] = {}

    for row in ws_signals.iter_rows(min_row=2, values_only=True):
        internal_name = str(row[0]).strip() if row[0] else ""
        if not internal_name:
            continue
        # Skip duplicates (keep first occurrence which has data)
        if internal_name in signals:
            continue
        unit = str(row[1]).strip() if row[1] else None
        description = str(row[2]).strip() if row[2] else None
        display_name_es = translate(description)
        display_name_en = description or DEFAULT_TRANSLATION
        signals[internal_name] = (display_name_es, display_name_en, unit)

    # ── Read equivalences ─────────────────────────
    ws_equiv = wb["equivalences"]
    # internal_name -> [(numeric_value, display_name_es, display_name_en)]
    equivalence_groups: list[tuple[str, list[tuple[float, str, str]]]] = []
    current_signal = None
    current_values: list[tuple[float, str, str]] = []

    for row in ws_equiv.iter_rows(min_row=2, values_only=True):
        sig = str(row[0]).strip() if row[0] else None
        val = row[1]
        name = str(row[2]).strip() if row[2] else None

        if sig is not None:
            # Save previous group
            if current_signal is not None and current_values:
                equivalence_groups.append((current_signal, current_values))
            current_signal = sig
            current_values = []

        if val is not None and name is not None:
            current_values.append((float(val), translate_equiv(name), name.strip()))

    # Save last group
    if current_signal is not None and current_values:
        equivalence_groups.append((current_signal, current_values))

    wb.close()

    # ── Generate Rust file ────────────────────────
    lines = [
        "// ════════════════════════════════════════════════════════════",
        "//  seed_data.rs — AUTO-GENERATED by scripts/generate_seed.py",
        f"//  Source: {EXCEL_PATH.name}",
        "//  Do not edit manually. Re-run `python scripts/generate_seed.py`",
        "// ════════════════════════════════════════════════════════════",
        "",
        "/// A signal definition from the seed catalog.",
        "#[derive(Debug, Clone, Copy)]",
        "pub(crate) struct SignalSeed {",
        "    pub internal_name: &'static str,",
        "    pub display_name_es: &'static str,",
        "    pub display_name_en: &'static str,",
        "    pub unit: Option<&'static str>,",
        "}",
        "",
        "/// A value mapping entry for a signal (numeric value \u2192 display name).",
        "#[derive(Debug, Clone, Copy)]",
        "pub(crate) struct ValueMappingSeed {",
        "    pub internal_name: &'static str,",
        "    pub numeric_value: f64,",
        "    pub display_name_es: &'static str,",
        "    pub display_name_en: &'static str,",
        "}",
        "",
        "/// All known OMNI signals with display names and units.",
        "pub(crate) static SIGNALS: &[SignalSeed] = &[",
    ]

    # Sort by internal_name for deterministic output
    sorted_signals = sorted(signals.items(), key=lambda x: x[0])
    for internal_name, (display_name_es, display_name_en, unit) in sorted_signals:
        unit_rust = to_rust_string(unit)
        lines.append(
            f'    SignalSeed {{ internal_name: "{internal_name}", '
            f'display_name_es: "{display_name_es}", '
            f'display_name_en: "{display_name_en}", '
            f'unit: {unit_rust} }},'
        )

    lines.extend([
        "];",
        "",
        "/// All known value mappings (state machine values \u2192 human-readable names).",
        "pub(crate) static VALUE_MAPPINGS: &[ValueMappingSeed] = &[",
    ])

    for signal_name, values in equivalence_groups:
        for numeric_value, display_name_es, display_name_en in values:
            lines.append(
                f'    ValueMappingSeed {{ internal_name: "{signal_name}", '
                f"numeric_value: {numeric_value}, "
                f'display_name_es: "{display_name_es}", '
                f'display_name_en: "{display_name_en}" }},'
            )

    lines.append("];")
    lines.append("")

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")

    # Stats
    n_equiv = sum(len(v) for _, v in equivalence_groups)
    n_groups = len(equivalence_groups)
    print(f"[OK] Generated {OUTPUT_PATH}")
    print(f"  Signals: {len(signals)} ({sum(1 for _, (_, _, u) in signals.items() if u is not None)} with unit)")
    print(f"  Value mappings: {n_equiv} entries across {n_groups} signal groups")
    untranslated = sum(1 for _, (d, _, _) in signals.items()
                       if d not in TRANSLATIONS.values() and d != DEFAULT_TRANSLATION)
    if untranslated:
        print(f"  WARNING: {untranslated} descriptions untranslated (using English fallback)")


if __name__ == "__main__":
    main()
